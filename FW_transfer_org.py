import pandas as pd
import openpyxl

# ファイルパス
file_aws = "/Users/yuu/Desktop/20250214_AWS_通信要件ヒアリングシート.xlsx"
file_internalfw = "/Users/yuu/Desktop/InternalFW_RuleList_Tokyo_Prod.xlsx"

# **[1] Internal FWシートのデータを取得**
wb_aws = openpyxl.load_workbook(file_aws, data_only=True)
ws_internal_fw = wb_aws["Internal FW"]

# AWS Account ID（D20のセル）を取得し、整数表記に修正
aws_account_id = str(int(ws_internal_fw["D20"].value))

# DataFrameとしてInternal FWのデータを取得
df_internal_fw = pd.DataFrame(ws_internal_fw.values)

# ヘッダー行の特定と適用
header_row_index = df_internal_fw[df_internal_fw.iloc[:, 7] == "Action"].index[0]
df_internal_fw.columns = df_internal_fw.iloc[header_row_index]
df_internal_fw = df_internal_fw[header_row_index + 1:].reset_index(drop=True)
df_internal_fw.columns = df_internal_fw.columns.str.strip()

# "追加/Add" の行を抽出
df_filtered = df_internal_fw[df_internal_fw["Action"].isin(["追加/Add"])].copy()

# **[2] Rules Prod シートのデータを取得**
wb_internalfw = openpyxl.load_workbook(file_internalfw)
ws_rules_prod = wb_internalfw["Rules Prod"]
ws_ip_variables = wb_internalfw["IP set variables Prod"]  # IP set variables Prodシートを取得

# IP set variablesの辞書を作成
ip_variables = {}
for row in ws_ip_variables.iter_rows(min_row=2, values_only=True):
    if row[1] and row[2]:  # B列（変数名）とC列（IP）が存在する場合
        # C列のIPアドレスを個別に保存
        ips = [ip.strip() for ip in str(row[2]).split('\n')]  # 改行で分割
        for ip in ips:
            if ip:  # 空でない場合
                ip_variables[ip.strip()] = row[1]  # 個々のIPアドレスと変数名を紐付け

def convert_to_variable(ip_str):
    # IPアドレスを抽出（カンマなしで分割）
    ips = [ip.strip() for ip in ip_str.replace('[', '').replace(']', '').split(',')]
    ips = [ip for ip in ips if ip]  # 空要素を削除
    
    # 各IPアドレスが変数テーブルに存在するか確認
    matched_variable = None
    for ip in ips:
        if ip in ip_variables:
            if matched_variable is None:
                matched_variable = ip_variables[ip]
            elif matched_variable != ip_variables[ip]:
                # 異なる変数に属するIPが混在している場合は元のIP形式を使用
                return f"[{','.join(ips)}]"
    
    # すべてのIPが同じ変数に属している場合は変数名を返す
    if matched_variable:
        return f"${matched_variable}"
    
    # 一致する変数がない場合は元のIP形式で返す
    return f"[{','.join(ips)}]"

# **[3] 転記データの作成**
df_filtered_expanded = pd.concat([
    df_filtered.assign(Action="alert"),
    df_filtered.assign(Action="pass")
], ignore_index=True)

# データ形式の調整
df_filtered_expanded["Protocol"] = df_filtered_expanded["Protocol"].str.lower()
df_filtered_expanded["Flow Option"] = df_filtered_expanded["Protocol"].apply(
    lambda x: "flow:to_server, established;" if x not in ["udp", "icmp"] else ""
)

# カラムの対応関係を設定
df_result = df_filtered_expanded.rename(columns={
    "Source IP address": "Source IP",
    "Destination IP address": "Destination IP",
    "Port number": "Destination Port"
})

# 必要なカラムのみを選択し、正しい順序で並べ替え
df_result = df_result[["Action", "Protocol", "Flow Option", "Source IP", "Destination IP", "Destination Port"]]

# **[4] alert, pass の順番を並び替え**
df_result = df_result.sort_values(by=["Source IP", "Destination IP", "Protocol"])
temp_groups = []
for _, group in df_result.groupby(["Source IP", "Destination IP", "Protocol"]):
    alert_rows = group[group["Action"] == "alert"]
    pass_rows = group[group["Action"] == "pass"]
    for a, p in zip(alert_rows.itertuples(index=False), pass_rows.itertuples(index=False)):
        temp_groups.extend([pd.DataFrame([a], columns=df_result.columns),
                          pd.DataFrame([p], columns=df_result.columns)])
df_result = pd.concat(temp_groups, ignore_index=True)

# 必要な列を追加（固定の開始値を使用）
df_result.insert(0, "項目", range(1, 1 + len(df_result)))  # 1から開始
df_result.insert(1, "sid", range(1000001, 1000001 + len(df_result)))  # 1000001から開始
df_result.insert(2, "履歴", pd.Timestamp.today().strftime("%Y/%m") + "/X追加")
df_result["Msg Option (AWS Account ID)"] = f'"{aws_account_id}"'

# Source IP / Destination IP の変換
df_result["Source IP"] = df_result["Source IP"].apply(convert_to_variable)
df_result["Destination IP"] = df_result["Destination IP"].apply(convert_to_variable)

# **[5] 新しいシートの作成と結果の出力**
# 新しいシートを作成（既に存在する場合は削除してから作成）
new_sheet_name = "Rules Prod New"
if new_sheet_name in wb_internalfw.sheetnames:
    wb_internalfw.remove(wb_internalfw[new_sheet_name])
ws_new = wb_internalfw.create_sheet(new_sheet_name)

# 元のシートからヘッダー行をコピー
for col_idx, cell in enumerate(ws_rules_prod[3], start=1):  # 3行目（ヘッダー行）をコピー
    ws_new.cell(row=3, column=col_idx, value=cell.value)
    # スタイルもコピー
    if cell.has_style:
        ws_new.cell(row=3, column=col_idx)._style = cell._style

# データを新しいシートに挿入
for idx, row in enumerate(df_result.itertuples(index=False), start=4):  # ヘッダー行の次から開始
    for col_idx, value in enumerate(row, start=1):
        ws_new.cell(row=idx, column=col_idx, value=value)
        # 元のシートと同じ列幅を設定
        if ws_rules_prod.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width:
            ws_new.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = \
                ws_rules_prod.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width

# シートの順序を調整（Rules Prodの次に配置）
wb_internalfw.move_sheet(new_sheet_name, offset=-len(wb_internalfw.sheetnames)+2)

# **[6] 保存**
output_file = "/Users/yuu/Desktop/InternalFW_RuleList_Tokyo_Prod_final.xlsx"
wb_internalfw.save(output_file)
